# restaurantes.py
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Form, Query
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Dict, Optional
from datetime import date
import os
import uuid
import openpyxl
from fpdf import FPDF

from app.db.database import get_db
from pydantic import EmailStr
from app.repositories.restaurante_repo import (
    get_all_restaurants,
    get_restaurant_stats,
    get_admins_by_restaurant,
    get_restaurant_by_id,
    get_restaurant_detail_stats,
    get_restaurant_sales,
    create_restaurante_with_admin,
)
from app.repositories.pedido_repo import get_pedidos_by_restaurante
from app.schemas.restaurante import (
    RestauranteOut,
    RestauranteDetailOut,
    RestauranteStatsOut,
    RestauranteStatsDetailOut,
    UsuarioOut,
)

from app.repositories.menu_repo import get_restaurant_menu
from app.schemas.menu import MenuRestauranteOut

router = APIRouter(
    prefix="/superadmin/restaurantes",
    tags=["SuperAdmin Restaurants"]
)

# ==========================
# DB DEPENDENCY
# ==========================


# ==========================
# LISTADO + BÚSQUEDA
# ==========================
@router.get("/", response_model=Dict)
async def list_restaurants(
    search: Optional[str] = None,
    estado: Optional[int] = Query(None),
    page: int = Query(1),
    limit: int = Query(12),
    db: AsyncSession = Depends(get_db)
):
    return await get_all_restaurants(db, search, estado, page, limit)


# ==========================
# DASHBOARD GLOBAL
# ==========================
@router.get("/stats", response_model=RestauranteStatsOut)
async def restaurant_stats(db: AsyncSession = Depends(get_db)):
    return await get_restaurant_stats(db)

@router.get("/sales", response_model=List[Dict])
async def sales_by_restaurant(db: AsyncSession = Depends(get_db)):
    return await get_restaurant_sales(db)

# ==========================
# DETALLE RESTAURANTE
# ==========================
@router.get("/{restaurante_id}", response_model=RestauranteDetailOut)
async def get_restaurant(
    restaurante_id: int,
    db: AsyncSession = Depends(get_db)
):
    restaurante = await get_restaurant_by_id(db, restaurante_id)
    if not restaurante:
        raise HTTPException(status_code=404, detail="Restaurante no encontrado")
    return restaurante

# ==========================
# STATS RESTAURANTE
# ==========================
@router.get("/{restaurante_id}/stats", response_model=RestauranteStatsDetailOut)
async def restaurant_detail_stats(
    restaurante_id: int,
    db: AsyncSession = Depends(get_db)
):
    return await get_restaurant_detail_stats(db, restaurante_id)

# ==========================
# ADMINS DEL RESTAURANTE
# ==========================
@router.get("/{restaurante_id}/admins", response_model=List[UsuarioOut])
async def admins_by_restaurant(
    restaurante_id: int,
    db: AsyncSession = Depends(get_db)
):
    return await get_admins_by_restaurant(db, restaurante_id)

# ==========================
# PEDIDOS RESTAURANTE
# ==========================
@router.get("/{restaurante_id}/pedidos")
async def pedidos_by_restaurant(
    restaurante_id: int,
    from_date: date | None = None,
    to_date: date | None = None,
    page: int = 1,
    limit: int = 20,
    db: AsyncSession = Depends(get_db),
):
    return await get_pedidos_by_restaurante(
        db,
        restaurante_id,
        from_date,
        to_date,
        page,
        limit,
    )

# ==========================
# CREAR RESTAURANTE
# ==========================
@router.post("/", response_model=RestauranteOut)
async def create_restaurant(
    nit: str = Form(...),
    nombre: str = Form(...),
    logo: UploadFile | None = File(None),
    admin_nombres: str = Form(...),
    admin_apellidos: str = Form(...),
    admin_documento: str = Form(...),
    admin_telefono: str = Form(...),
    admin_correo: EmailStr = Form(...),
    admin_contrasena: str = Form(...),
    admin_tipo_documento_id: int = Form(...),
    db: AsyncSession = Depends(get_db),
):
    logo_path = None
    if logo:
        upload_dir = "uploads/logos"
        os.makedirs(upload_dir, exist_ok=True)
        ext = os.path.splitext(logo.filename)[1]
        file_name = f"{uuid.uuid4()}{ext}"
        file_path = os.path.join(upload_dir, file_name)
        with open(file_path, "wb") as f:
            f.write(await logo.read())
        logo_path = f"/uploads/logos/{file_name}"

    restaurante_data = {
        "nit": nit,
        "nombre": nombre,
        "logo": logo_path,
        "admin_nombres": admin_nombres,
        "admin_apellidos": admin_apellidos,
        "admin_documento": admin_documento,
        "admin_telefono": admin_telefono,
        "admin_correo": admin_correo,
        "admin_contrasena": admin_contrasena,
        "admin_tipo_documento_id": admin_tipo_documento_id,
    }

    restaurante = await create_restaurante_with_admin(db, restaurante_data)
    return restaurante

# ==========================
# REPORTES PEDIDOS
# ==========================
@router.get("/{restaurante_id}/reportes/pedidos")
async def download_pedidos(
    restaurante_id: int,
    from_date: date = Query(...),
    to_date: date = Query(...),
    tipo: str = Query("excel"),  # "excel" o "pdf"
    db: AsyncSession = Depends(get_db)
):

    restaurante = await get_restaurant_by_id(db, restaurante_id)
    if not restaurante:
        raise HTTPException(status_code=404, detail="Restaurante no encontrado")

    report_dir = "reportes"
    os.makedirs(report_dir, exist_ok=True)
    file_name = f"pedidos_{restaurante_id}.{ 'xlsx' if tipo=='excel' else 'pdf' }"
    file_path = os.path.join(report_dir, file_name)

    pedidos_data = await get_pedidos_by_restaurante(
        db, restaurante_id, from_date, to_date, page=1, limit=1000
    )
    pedidos_list = pedidos_data["items"]

    if tipo == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID Pedido", "Cliente", "Total", "Estado", "Fecha"])
        for p in pedidos_list:
            ws.append([
                p.get("id_pedido"),
                p.get("cliente"),
                p.get("total"),
                p.get("estado"),
                p.get("fecha").strftime("%Y-%m-%d") if p.get("fecha") else "",
            ])
        wb.save(file_path)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(0, 10, f"Pedidos Restaurante {restaurante.nombre}", ln=True)
        for p in pedidos_list:
            pdf.cell(0, 10, f"{p['id_pedido']} | {p['cliente']} | {p['total']} | {p['estado']} | {p['fecha']}", ln=True)
        pdf.output(file_path)
        media_type = "application/pdf"

    return FileResponse(path=file_path, filename=file_name, media_type=media_type)


# ==========================
# REPORTES RESUMEN (solo Excel)
# ==========================
@router.get("/{restaurante_id}/reportes/resumen")
async def download_resumen(
    restaurante_id: int,
    tipo: str = Query("excel"),
    db: AsyncSession = Depends(get_db)
):
    if tipo != "excel":
        raise HTTPException(status_code=400, detail="Solo disponible en Excel")

    restaurante = await get_restaurant_by_id(db, restaurante_id)
    if not restaurante:
        raise HTTPException(status_code=404, detail="Restaurante no encontrado")

    report_dir = "reportes"
    os.makedirs(report_dir, exist_ok=True)
    file_name = f"resumen_{restaurante_id}.xlsx"
    file_path = os.path.join(report_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Total pedidos", "Total vendido"])
    stats = await get_restaurant_detail_stats(db, restaurante_id)
    ws.append([stats["total_pedidos"], stats["total_vendido"]])
    wb.save(file_path)

    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(path=file_path, filename=file_name, media_type=media_type)


# ==========================
# ACTUALIZAR RESTAURANTE (estado / logo)
# ==========================
@router.patch("/{restaurante_id}")
async def update_restaurante(
    restaurante_id: int,
    estado_id: int | None = Form(None),
    logo: UploadFile | None = File(None),
    db: AsyncSession = Depends(get_db)
):
    from app.models.restaurante import Restaurante
    from sqlalchemy import update, select

    # Verificar que el restaurante exista
    query = select(Restaurante).where(Restaurante.id_restaurante == restaurante_id)
    result = await db.execute(query)
    restaurante = result.scalars().first()
    if not restaurante:
        raise HTTPException(status_code=404, detail="Restaurante no encontrado")

    # Preparar datos de actualización
    update_data = {}
    if estado_id is not None:
        update_data["estado_id"] = estado_id

    if logo:
        upload_dir = "uploads/logos"
        os.makedirs(upload_dir, exist_ok=True)
        ext = os.path.splitext(logo.filename)[1]
        file_name = f"{uuid.uuid4()}{ext}"
        file_path = os.path.join(upload_dir, file_name)
        with open(file_path, "wb") as f:
            f.write(await logo.read())
        update_data["logo"] = f"/uploads/logos/{file_name}"

    # Ejecutar update
    if update_data:
        stmt = (
            update(Restaurante)
            .where(Restaurante.id_restaurante == restaurante_id)
            .values(**update_data)
        )
        await db.execute(stmt)
        await db.commit()

    # Volver a obtener el restaurante actualizado
    query = select(Restaurante).where(Restaurante.id_restaurante == restaurante_id)
    result = await db.execute(query)
    restaurante = result.scalars().first()

    return {
        "id_restaurante": restaurante.id_restaurante,
        "nombre": restaurante.nombre,
        "logo": restaurante.logo,
        "estado_id": restaurante.estado_id,
        "fecha_creacion": restaurante.fecha_creacion.isoformat(),
    }

@router.get(
    "/{id}/menu",
    response_model=MenuRestauranteOut
)
async def get_menu_restaurante(
    id: int,
    db: AsyncSession = Depends(get_db)
):
    categorias = await get_restaurant_menu(db, id)
    return {"categorias": categorias}